from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl import load_workbook
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)

STOCK_FILE = "stock.xlsx"
VENTAS_FILE = "ventas.xlsx"


# ==========================
# Crear ventas.xlsx si no existe
# ==========================

def crear_ventas():

    if not os.path.exists(VENTAS_FILE):

        from openpyxl import Workbook

        wb = Workbook()

        ws = wb.active

        ws.title = "Ventas"

        ws.append([
            "Fecha",
            "Producto",
            "Cantidad",
            "Precio Unitario",
            "Total"
        ])

        wb.save(VENTAS_FILE)


crear_ventas()


# ==========================
# Leer catálogo
# ==========================

def obtener_catalogo():

    wb = load_workbook(STOCK_FILE)

    ws = wb.active

    mangas = []

    for fila in ws.iter_rows(min_row=2, values_only=True):

        mangas.append({

            "id": fila[0],
            "nombre": fila[1],
            "autor": fila[2],
            "precio": fila[3],
            "stock": fila[4],
            "imagen": fila[5],
            "categoria": fila[6]

        })

    wb.close()

    return mangas


# ==========================
# API
# ==========================

@app.route("/catalogo", methods=["GET"])
def catalogo():

    return jsonify(obtener_catalogo())


# ==========================
# Comprar
# ==========================

@app.route("/comprar", methods=["POST"])
def comprar():

    datos = request.get_json()

    id_manga = datos["id"]

    cantidad = datos["cantidad"]

    wb = load_workbook(STOCK_FILE)

    ws = wb.active

    for fila in range(2, ws.max_row + 1):

        if ws[fila][0].value == id_manga:

            stock = ws[fila][4].value

            if stock < cantidad:

                wb.close()

                return jsonify({

                    "success": False,
                    "mensaje": "No hay suficiente stock."

                }), 400

            ws[fila][4].value = stock - cantidad

            precio = ws[fila][3].value

            nombre = ws[fila][1].value

            wb.save(STOCK_FILE)

            wb.close()

            registrar_venta(

                nombre,
                cantidad,
                precio

            )

            return jsonify({

                "success": True,
                "mensaje": "Compra realizada correctamente."

            })

    wb.close()

    return jsonify({

        "success": False,
        "mensaje": "Producto no encontrado."

    }), 404


# ==========================
# Registrar venta
# ==========================

def registrar_venta(

        producto,
        cantidad,
        precio

):

    wb = load_workbook(VENTAS_FILE)

    ws = wb.active

    ws.append([

        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),

        producto,

        cantidad,

        precio,

        cantidad * precio

    ])

    wb.save(VENTAS_FILE)

    wb.close()


# ==========================
# Restablecer inventario
# ==========================

@app.route("/reiniciar", methods=["POST"])
def reiniciar():

    wb = load_workbook(STOCK_FILE)

    ws = wb.active

    for fila in range(2, ws.max_row + 1):

        ws[fila][4].value = 100

    wb.save(STOCK_FILE)

    wb.close()

    return jsonify({

        "mensaje": "Inventario reiniciado."

    })


# ==========================
# Stock disponible
# ==========================

@app.route("/stock", methods=["GET"])
def stock():

    return jsonify(obtener_catalogo())


# ==========================
# Historial de ventas
# ==========================

@app.route("/ventas", methods=["GET"])
def ventas():

    wb = load_workbook(VENTAS_FILE)

    ws = wb.active

    historial = []

    for fila in ws.iter_rows(min_row=2, values_only=True):

        historial.append({

            "fecha": fila[0],
            "producto": fila[1],
            "cantidad": fila[2],
            "precio": fila[3],
            "total": fila[4]

        })

    wb.close()

    return jsonify(historial)


# ==========================
# Página principal
# ==========================

@app.route("/")
def inicio():

    return jsonify({

        "mensaje": "Servidor Dark Mangas funcionando correctamente."

    })


# ==========================
# Ejecutar servidor
# ==========================

if __name__ == "__main__":

    app.run(

        debug=True,

        host="0.0.0.0",

        port=5000

    )